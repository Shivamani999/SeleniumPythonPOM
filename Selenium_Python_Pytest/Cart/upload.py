import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait


def update_excel(file_path, row_name, col_name, value):
    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    Dict = {}

    for i in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=i).value == col_name:
            Dict["col"] = i

    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            if sheet.cell(row=i, column=j).value == row_name:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = value
    book.save(file_path)

driver = webdriver.Chrome()
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.implicitly_wait(15)
driver.maximize_window()

file_path = "C:\\Users\\lenovo\\Downloads\\download.xlsx"

update_excel(file_path, "Apple", "price", "990")

file_upload = driver.find_element(By.ID, "fileinput")
file_upload.send_keys(file_path)

toast = (By.XPATH, "//div[@class='Toastify__toast-body']/div[2]")
WebDriverWait(driver, 10).until(expected_conditions.visibility_of_element_located(toast))
print(driver.find_element(*toast).text)
