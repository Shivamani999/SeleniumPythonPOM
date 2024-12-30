import openpyxl


class homepageData:

    # user_creds = [{"firstname":"Shivamani","lastname":"Konam","gender":"Male"}, {"firstname":"Harshitha","lastname":"Konam","gender":"Female"}]

    def getTestData(testcasename):
        Dict = {}
        book = openpyxl.load_workbook("D:\\Excel.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testcasename:
                for j in range(1, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
